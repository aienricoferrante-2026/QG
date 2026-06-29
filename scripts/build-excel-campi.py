#!/usr/bin/env python3
# Costruisce CATALOGO_CAMPI_ERP.xlsx: tutti i campi del DB unico (bqyqr),
# dove stanno, a cosa servono, e se prendono da Qnet.
import json, os, re, urllib.request, urllib.error
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
ENV=os.path.expanduser("~/Desktop/qualifica-platform/apps/hub/.env")
TOK=next((re.match(r'\s*ACCESS_TOKEN_ACCOUNT\s*=\s*["\']?([^"\'\r\n]+)',l).group(1).strip() for l in open(ENV) if l.startswith("ACCESS_TOKEN_ACCOUNT")),None)
HUB="bqyqrqmbekdhejrzasvv"
def q(sql):
    req=urllib.request.Request(f"https://api.supabase.com/v1/projects/{HUB}/database/query",data=json.dumps({"query":sql}).encode(),headers={"Authorization":f"Bearer {TOK}","Content-Type":"application/json","User-Agent":"curl/8"})
    with urllib.request.urlopen(req,timeout=180) as r: return json.load(r)
SCHEMAS=['public','commerciale','commesse','formazione','sedi_partner','contabilita','cdg','hr','iso','sic','fia','bp']
DOMINIO={'public':'Anagrafica master','commerciale':'Commerciale/CRM','commesse':'Commesse','formazione':'Formazione (FOR)','sedi_partner':'Sedi & Partner','contabilita':'Contabilità (attiva+passiva)','cdg':'Controllo di gestione','hr':'HR / Personale','iso':'ISO / Qualità','sic':'Sicurezza','fia':'Bandi & Incentivi (FIA)','bp':'Business Plan'}
inlist="("+",".join(f"'{s}'" for s in SCHEMAS)+")"
# colonne (per schema, evita payload enormi)
cols=[]
for s in SCHEMAS:
    r=q(f"""select c.table_schema sch, c.table_name tbl, c.column_name col, c.ordinal_position ord,
        coalesce(c.udt_name,c.data_type) typ, c.is_nullable nul, c.column_default def,
        col_description(format('%s.%s',c.table_schema,c.table_name)::regclass, c.ordinal_position) cmt
        from information_schema.columns c
        where c.table_schema='{s}' and c.table_name not like 'v\\_%' order by c.table_name, c.ordinal_position""")
    cols+=r
# PK + FK
pk=set((r['sch'],r['tbl'],r['col']) for r in q(f"""select n.nspname sch, t.relname tbl, a.attname col from pg_index i join pg_class t on t.oid=i.indrelid join pg_namespace n on n.oid=t.relnamespace join pg_attribute a on a.attrelid=t.oid and a.attnum=any(i.indkey) where i.indisprimary and n.nspname in {inlist}"""))
fk={}
for r in q(f"""select n.nspname sch, c.relname tbl, a.attname col, fn.nspname fsch, ft.relname ftbl
    from pg_constraint con join pg_class c on c.oid=con.conrelid join pg_namespace n on n.oid=c.relnamespace
    join pg_class ft on ft.oid=con.confrelid join pg_namespace fn on fn.oid=ft.relnamespace
    join pg_attribute a on a.attrelid=c.oid and a.attnum=any(con.conkey)
    where con.contype='f' and n.nspname in {inlist}"""):
    fk[(r['sch'],r['tbl'],r['col'])]=f"{r['fsch']}.{r['ftbl']}"

SYS={'id','created_at','updated_at','created_by','created_by_utente_id','updated_by','deleted_at','archiviato_il','hash_record','hash_prev'}
def fonte(c):
    n=c['col'].lower()
    if 'qnet' in n or 'qnet' in (c['tbl'].lower()): return 'Qnet'
    if c['cmt'] and 'qnet' in c['cmt'].lower(): return 'Qnet'
    if n in SYS or n.endswith('_at') or n.startswith('hash_'): return 'Sistema'
    return 'Interno/App'
def descr(c):
    if c['cmt']: return c['cmt']
    n=c['col'].lower(); k=(c['sch'],c['tbl'],c['col'])
    if n=='qnet_id': return "ID della riga in Qnet (chiave di sincronizzazione)"
    if n.endswith('_qnet_id'): return f"ID Qnet di «{n[:-8]}» (sincronizzazione)"
    if n in('codice',): return "Codice identificativo"
    if k in fk: return f"Riferimento (FK) a {fk[k]}"
    if n=='id': return "Identificativo univoco (PK)"
    if n.endswith('_id'): return f"Riferimento a «{n[:-3]}»"
    if n in('created_at','updated_at'): return "Data/ora di "+("creazione" if 'created' in n else "ultima modifica")+" (sistema)"
    if n.endswith('_at'): return "Data/ora di "+n[:-3].replace('_',' ')
    if n.startswith('is_') or n.startswith('has_') or n in('attivo','archiviato'): return f"Flag sì/no: {n.replace('is_','').replace('has_','').replace('_',' ')}"
    if n.startswith('data_') or n.startswith('data'): return f"Data: {n.replace('data_','').replace('_',' ')}"
    if 'importo' in n or 'valore' in n or 'ricav' in n or 'costo' in n or 'mol' in n or 'totale' in n or 'pct' in n or 'percentuale' in n: return f"Valore economico/numerico: {n.replace('_',' ')}"
    if n in('codice','code') or n.endswith('_codice'): return f"Codice: {n.replace('_',' ')}"
    if 'nome' in n or 'titolo' in n or 'descrizione' in n or 'note' in n or 'ragione' in n: return f"Testo: {n.replace('_',' ')}"
    if 'email' in n: return "Email"
    if 'stato' in n or 'status' in n: return f"Stato: {n.replace('_',' ')}"
    return c['col'].replace('_',' ').capitalize()

# === xlsx ===
wb=Workbook(); HEAD=Font(bold=True,color="FFFFFF"); FILL=PatternFill("solid",fgColor="1F4E78")
QF=PatternFill("solid",fgColor="FCE4D6"); SF=PatternFill("solid",fgColor="EDEDED")
thin=Side(style='thin',color="D9D9D9"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
# Sheet LEGENDA
ws=wb.active; ws.title="LEGENDA"
ws["A1"]="CATALOGO CAMPI — ERP Qualifica (database unico bqyqr)"; ws["A1"].font=Font(bold=True,size=14)
leg=[("",""),("Totale campi",str(len(cols))),("Domini",str(len(SCHEMAS))),("",""),
 ("Colonna","Significato"),("Dominio","Area di business (schema del DB)"),("Tabella","Tabella che contiene il campo"),
 ("Campo","Nome del campo (colonna)"),("Tipo","Tipo di dato"),("Obblig.","Sì = obbligatorio (NOT NULL)"),
 ("PK","● = chiave primaria"),("FK →","Se è un riferimento, la tabella puntata"),
 ("A cosa serve","Descrizione (dal commento DB o dedotta dal nome)"),
 ("Fonte","Qnet = proviene/sincronizza da Qnet · Sistema = generato automaticamente (id/date/hash) · Interno/App = gestito dall'app/utente"),
 ("",""),("NOTA Fonte","Il flag Qnet è basato su naming (campi/tabelle 'qnet') + commenti. Per la mappa Qnet↔campo di dettaglio: REGISTRO_PARITA_QNET_WEA.md (246 campi) e packages/qualifica-qnet/API-REFERENCE.md (643 campi Qnet).")]
for i,(a,b) in enumerate(leg,3):
    ws[f"A{i}"]=a; ws[f"B{i}"]=b; ws[f"A{i}"].font=Font(bold=True)
ws.column_dimensions['A'].width=22; ws.column_dimensions['B'].width=110
# Sheet INDICE (tabelle per dominio)
wi=wb.create_sheet("INDICE")
wi.append(["Dominio","Schema","Tabella","N° campi","di cui Qnet"]);
for c in wi[1]: c.font=HEAD; c.fill=FILL
from collections import defaultdict
bytbl=defaultdict(list)
for c in cols: bytbl[(c['sch'],c['tbl'])].append(c)
for (s,t),cc in sorted(bytbl.items()):
    nq=sum(1 for c in cc if fonte(c)=='Qnet')
    wi.append([DOMINIO.get(s,s),s,t,len(cc),nq])
for col in 'ABCDE': wi.column_dimensions[col].width=[26,16,40,10,12]['ABCDE'.index(col)]
wi.freeze_panes="A2"
# Sheet CAMPI (tutti)
wc=wb.create_sheet("CAMPI");
HDR=["Dominio","Schema","Tabella","Campo","Tipo","Obblig.","PK","FK →","A cosa serve","Fonte"]
wc.append(HDR)
for c in wc[1]: c.font=HEAD; c.fill=FILL; c.alignment=Alignment(vertical="center")
for c in cols:
    k=(c['sch'],c['tbl'],c['col']); f=fonte(c)
    row=[DOMINIO.get(c['sch'],c['sch']),c['sch'],c['tbl'],c['col'],c['typ'],
         "Sì" if c['nul']=='NO' else "",'●' if k in pk else "",fk.get(k,""),descr(c),f]
    wc.append(row)
    r=wc.max_row
    if f=='Qnet':
        for cc in range(1,11): wc.cell(r,cc).fill=QF
    elif f=='Sistema':
        for cc in range(1,11): wc.cell(r,cc).fill=SF
widths=[24,14,30,28,16,8,5,26,55,13]
for i,w in enumerate(widths,1): wc.column_dimensions[get_column_letter(i)].width=w
wc.freeze_panes="A2"; wc.auto_filter.ref=f"A1:J{wc.max_row}"
out=os.path.expanduser("~/Desktop/STW/CATALOGO_CAMPI_ERP.xlsx")
wb.save(out)
nq=sum(1 for c in cols if fonte(c)=='Qnet'); ns=sum(1 for c in cols if fonte(c)=='Sistema')
print(f"OK {out}\n  {len(cols)} campi · {len(bytbl)} tabelle · Qnet {nq} · Sistema {ns} · Interno {len(cols)-nq-ns}")
