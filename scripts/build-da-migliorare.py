#!/usr/bin/env python3
# Aggiunge il foglio "DA MIGLIORARE" a CATALOGO_CAMPI_ERP.xlsx: anomalie del modello.
import json, os, re, urllib.request
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
ENV=os.path.expanduser("~/Desktop/qualifica-platform/apps/hub/.env")
TOK=next((re.match(r'\s*ACCESS_TOKEN_ACCOUNT\s*=\s*["\']?([^"\'\r\n]+)',l).group(1).strip() for l in open(ENV) if l.startswith("ACCESS_TOKEN_ACCOUNT")),None)
HUB="bqyqrqmbekdhejrzasvv"
def q(sql):
    req=urllib.request.Request(f"https://api.supabase.com/v1/projects/{HUB}/database/query",data=json.dumps({"query":sql}).encode(),headers={"Authorization":f"Bearer {TOK}","Content-Type":"application/json","User-Agent":"curl/8"})
    with urllib.request.urlopen(req,timeout=180) as r: return json.load(r)
SCH="('public','commerciale','commesse','formazione','sedi_partner','contabilita','cdg','hr','iso','sic','fia','bp')"
# 1) tabelle vuote (conteggio reale)
vuote=q(f"""do $$ begin end $$; select n.nspname sch, c.relname tbl,
  (xpath('/r/c/text()', query_to_xml(format('select count(*) c from %I.%I',n.nspname,c.relname),false,true,'')))[1]::text::bigint righe
  from pg_class c join pg_namespace n on n.oid=c.relnamespace
  where n.nspname in {SCH} and c.relkind='r' order by 1,2""")
empties=[(r['sch'],r['tbl']) for r in vuote if r['righe']==0]
# 2) tabelle senza PK
nopk=q(f"select n.nspname sch, c.relname tbl from pg_class c join pg_namespace n on n.oid=c.relnamespace where n.nspname in {SCH} and c.relkind='r' and not exists (select 1 from pg_index i where i.indrelid=c.oid and i.indisprimary) order by 1,2")
# 3) colonne *_id senza FK
nofk=q(f"""select col.table_schema sch, col.table_name tbl, col.column_name col from information_schema.columns col
  where col.table_schema in {SCH} and col.column_name like '%\\_id' and col.column_name<>'qnet_id'
  and not exists (select 1 from information_schema.key_column_usage k join information_schema.table_constraints t on t.constraint_name=k.constraint_name and t.constraint_type='FOREIGN KEY' where k.table_schema=col.table_schema and k.table_name=col.table_name and k.column_name=col.column_name)
  order by 1,2,3""")
# 4) qnet_id non agganciati (tabelle con qnet_id, % NULL)
qn=q(f"""select c.table_schema sch, c.table_name tbl from information_schema.columns c join information_schema.tables t on t.table_schema=c.table_schema and t.table_name=c.table_name and t.table_type='BASE TABLE' where c.table_schema in {SCH} and c.column_name='qnet_id'""")
qnull=[]
for r in qn:
    try:
        d=q(f"select count(*) tot, count(*) filter (where qnet_id is null) nulli from {r['sch']}.{r['tbl']}")[0]
        if d['tot']>0 and d['nulli']>0: qnull.append((r['sch'],r['tbl'],d['tot'],d['nulli'],round(100*d['nulli']/d['tot'])))
    except Exception: pass

wb=load_workbook(os.path.expanduser("~/Desktop/STW/CATALOGO_CAMPI_ERP.xlsx"))
if "DA MIGLIORARE" in wb.sheetnames: del wb["DA MIGLIORARE"]
ws=wb.create_sheet("DA MIGLIORARE", 2)
H=Font(bold=True,color="FFFFFF"); F=PatternFill("solid",fgColor="C00000"); SUB=Font(bold=True,size=12,color="C00000")
def hdr(row):
    for c in row: c.font=H; c.fill=F
r=1
ws.cell(r,1,"DA MIGLIORARE — anomalie del modello dati (candidati per le migliorie)").font=Font(bold=True,size=14); r+=1
ws.cell(r,1,"NB: sono CANDIDATI DA RIVEDERE, non tutti difetti. Alcune tabelle senza PK sono aggregati BI (es. cdg.conto_periodo); alcuni qnet-null sono per-disegno (sedi, opportunità interne).").font=Font(italic=True,size=9); r+=2
ws.cell(r,1,f"① TABELLE VUOTE ({len(empties)}) — scaffold/non usate: o si popolano, o si droppano").font=SUB; r+=1
ws.cell(r,1,"Dominio").font=H; ws.cell(r,2,"Tabella").font=H
for c in (ws.cell(r,1),ws.cell(r,2)): c.fill=F
r+=1
for s,t in empties: ws.cell(r,1,s); ws.cell(r,2,t); r+=1
r+=1
ws.cell(r,1,f"② TABELLE SENZA CHIAVE PRIMARIA ({len(nopk)}) — rischio duplicati/integrità").font=SUB; r+=1
ws.cell(r,1,"Dominio").font=H; ws.cell(r,2,"Tabella").font=H
for c in (ws.cell(r,1),ws.cell(r,2)): c.fill=F
r+=1
for x in nopk: ws.cell(r,1,x['sch']); ws.cell(r,2,x['tbl']); r+=1
r+=1
ws.cell(r,1,f"③ RIFERIMENTI *_id SENZA FK ({len(nofk)}) — link non garantiti dal DB (rischio orfani)").font=SUB; r+=1
for i,h in enumerate(["Dominio","Tabella","Campo"],1): c=ws.cell(r,i,h); c.font=H; c.fill=F
r+=1
for x in nofk: ws.cell(r,1,x['sch']); ws.cell(r,2,x['tbl']); ws.cell(r,3,x['col']); r+=1
r+=1
ws.cell(r,1,f"④ QNET NON AGGANCIATI ({len(qnull)} tabelle) — righe senza qnet_id (non matchate/sincronizzate)").font=SUB; r+=1
for i,h in enumerate(["Dominio","Tabella","Righe tot","Senza qnet_id","% non agganciato"],1): c=ws.cell(r,i,h); c.font=H; c.fill=F
r+=1
for s,t,tot,nu,pct in sorted(qnull,key=lambda x:-x[4]):
    ws.cell(r,1,s); ws.cell(r,2,t); ws.cell(r,3,tot); ws.cell(r,4,nu); ws.cell(r,5,f"{pct}%"); r+=1
for i,w in enumerate([24,34,14,16,18],1): ws.column_dimensions[get_column_letter(i)].width=w
wb.save(os.path.expanduser("~/Desktop/STW/CATALOGO_CAMPI_ERP.xlsx"))
print(f"DA MIGLIORARE aggiunto: vuote={len(empties)} · senza-PK={len(nopk)} · *_id-senza-FK={len(nofk)} · qnet-non-agganciati={len(qnull)} tab")
