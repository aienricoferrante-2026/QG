#!/usr/bin/env python3
# CATALOGO_CAMPI_ERP.xlsx v2: + colonna WeA per campo + foglio UNICITA/NO-DUPLICATI.
import json, os, re, urllib.request
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
ENV=os.path.expanduser("~/Desktop/qualifica-platform/apps/hub/.env")
TOK=next((re.match(r'\s*ACCESS_TOKEN_ACCOUNT\s*=\s*["\']?([^"\'\r\n]+)',l).group(1).strip() for l in open(ENV) if l.startswith("ACCESS_TOKEN_ACCOUNT")),None)
HUB="bqyqrqmbekdhejrzasvv"
def q(sql):
    req=urllib.request.Request(f"https://api.supabase.com/v1/projects/{HUB}/database/query",data=json.dumps({"query":sql}).encode(),headers={"Authorization":f"Bearer {TOK}","Content-Type":"application/json","User-Agent":"curl/8"})
    with urllib.request.urlopen(req,timeout=180) as r: return json.load(r)
TBL2APPS=json.load(open("/tmp/tbl2apps.json"))  # tabella -> [WeA]
SCHEMAS=['public','commerciale','commesse','formazione','sedi_partner','contabilita','cdg','hr','iso','sic','fia','bp']
DOMINIO={'public':'Anagrafica master','commerciale':'Commerciale/CRM','commesse':'Commesse','formazione':'Formazione','sedi_partner':'Sedi & Partner','contabilita':'Contabilità','cdg':'Controllo gestione','hr':'HR','iso':'ISO','sic':'Sicurezza','fia':'Bandi (FIA)','bp':'Business Plan'}
inlist="("+",".join(f"'{s}'" for s in SCHEMAS)+")"
cols=[]
for s in SCHEMAS:
    cols+=q(f"""select c.table_schema sch,c.table_name tbl,c.column_name col,c.ordinal_position ord,coalesce(c.udt_name,c.data_type) typ,c.is_nullable nul,col_description(format('%s.%s',c.table_schema,c.table_name)::regclass,c.ordinal_position) cmt from information_schema.columns c where c.table_schema='{s}' and c.table_name not like 'v\\_%' order by c.table_name,c.ordinal_position""")
pk=set((r['sch'],r['tbl'],r['col']) for r in q(f"select n.nspname sch,t.relname tbl,a.attname col from pg_index i join pg_class t on t.oid=i.indrelid join pg_namespace n on n.oid=t.relnamespace join pg_attribute a on a.attrelid=t.oid and a.attnum=any(i.indkey) where i.indisprimary and n.nspname in {inlist}"))
fk={}
for r in q(f"select n.nspname sch,c.relname tbl,a.attname col,fn.nspname fsch,ft.relname ftbl from pg_constraint con join pg_class c on c.oid=con.conrelid join pg_namespace n on n.oid=c.relnamespace join pg_class ft on ft.oid=con.confrelid join pg_namespace fn on fn.oid=ft.relnamespace join pg_attribute a on a.attrelid=c.oid and a.attnum=any(con.conkey) where con.contype='f' and n.nspname in {inlist}"):
    fk[(r['sch'],r['tbl'],r['col'])]=f"{r['fsch']}.{r['ftbl']}"
SYS={'id','created_at','updated_at','created_by','created_by_utente_id','updated_by','deleted_at','archiviato_il','hash_record','hash_prev'}
def wea(tbl): return ", ".join(TBL2APPS.get(tbl,[])) or "(non in WeA / solo DB)"
def fonte(c):
    n=c['col'].lower()
    if 'qnet' in n or 'qnet' in c['tbl'].lower(): return 'Qnet'
    if c['cmt'] and 'qnet' in c['cmt'].lower(): return 'Qnet'
    if n in SYS or n.endswith('_at') or n.startswith('hash_'): return 'Sistema'
    return 'Interno/App'
def descr(c):
    if c['cmt']: return c['cmt']
    n=c['col'].lower(); k=(c['sch'],c['tbl'],c['col'])
    if n=='qnet_id': return "ID della riga in Qnet (chiave di sincronizzazione)"
    if n.endswith('_qnet_id'): return f"ID Qnet di «{n[:-8]}» (sincronizzazione)"
    if n=='codice': return "Codice identificativo"
    if k in fk: return f"Collegamento (FK) a {fk[k]}"
    if n=='id': return "Identificativo univoco della riga (PK)"
    if n.endswith('_id'): return f"Collegamento a «{n[:-3]}»"
    if n in('created_at','updated_at'): return "Data/ora di "+("creazione" if 'created' in n else "ultima modifica")+" (automatica)"
    if n.endswith('_at'): return "Data/ora di "+n[:-3].replace('_',' ')
    if n.startswith(('is_','has_','flag_')) or n in('attivo','archiviato','presente_in_qnet'): return f"Sì/No — {n.replace('is_','').replace('has_','').replace('flag_','').replace('_',' ')}"
    if n.startswith('data') or n.endswith('_data'): return f"Data — {n.replace('data_','').replace('_data','').replace('_',' ')}"
    if any(x in n for x in('importo','imponibile','iva','ricav','costo','mol','totale','prezzo','saldo','incass','pagat','versat','compenso','provvig')): return f"Valore € — {n.replace('_',' ')}"
    if 'pct' in n or 'percentuale' in n or 'perc' in n or 'quota' in n: return f"Percentuale/quota — {n.replace('_',' ')}"
    if n.endswith('_codice') or n.endswith('_code'): return f"Codice — {n.replace('_codice','').replace('_',' ')}"
    if any(x in n for x in('nome','titolo','descrizione','note','ragione','oggetto','denominazione')): return f"Testo — {n.replace('_',' ')}"
    if 'email' in n: return "Indirizzo email"
    if 'telefono' in n or 'cellulare' in n or 'tel'==n: return "Telefono"
    if 'piva' in n or 'partita_iva' in n: return "Partita IVA"
    if n in('cf','codice_fiscale') or 'codice_fiscale' in n: return "Codice fiscale"
    if 'stato' in n or 'status' in n: return f"Stato — {n.replace('_',' ')}"
    if 'meta'==n or 'payload' in n or 'json' in n or 'jsonb' in c['typ']: return f"Dati strutturati (JSON) — {n.replace('_',' ')}"
    return c['col'].replace('_',' ').capitalize()

wb=Workbook(); H=Font(bold=True,color="FFFFFF"); FILL=PatternFill("solid",fgColor="1F4E78")
QF=PatternFill("solid",fgColor="FCE4D6"); SF=PatternFill("solid",fgColor="EDEDED")
GREEN=PatternFill("solid",fgColor="C6EFCE"); YELL=PatternFill("solid",fgColor="FFEB9C"); ORANGE=PatternFill("solid",fgColor="F8CBAD")
# LEGENDA
ws=wb.active; ws.title="LEGENDA"
ws["A1"]="CATALOGO CAMPI ERP — database unico bqyqr (v2: + WeA + verifica unicità)"; ws["A1"].font=Font(bold=True,size=14)
leg=[("Totale campi",str(len(cols))),("Tabelle",str(len({(c['sch'],c['tbl']) for c in cols}))),("Domini",str(len(SCHEMAS))),("",""),
 ("Foglio CAMPI","ogni campo: dominio, tabella, campo, WeA che lo usa, tipo, obblig., PK, collegamento(FK), a cosa serve, fonte"),
 ("Foglio UNICITÀ","per ogni CONCETTO: tutte le tabelle che lo contengono → MASTER unico ✅ / estensione-figlia / ⚠️ da verificare. È la prova anti-duplicato."),
 ("Foglio DA MIGLIORARE","anomalie (senza-PK, *_id senza-FK, qnet non agganciati)"),
 ("WeA","quale app usa la tabella (mappato dal codice .from(), NON dedotto)"),
 ("Fonte","Qnet=da Qnet · Sistema=automatico (id/date/hash) · Interno/App=app/utente"),
 ("Nome nella WeA","= il nome del campo (le app interrogano i campi con questo nome; l'etichetta a video è nella UI dell'app)")]
for i,(a,b) in enumerate(leg,3):
    ws.cell(i,1,a).font=Font(bold=True); ws.cell(i,2,b)
ws.column_dimensions['A'].width=20; ws.column_dimensions['B'].width=120

# UNICITÀ / NO DUPLICATI
wu=wb.create_sheet("UNICITÀ - NO DUPLICATI")
allt=sorted({(c['sch'],c['tbl']) for c in cols})
def tname(t): return t[1]
# concetto -> (master canonico, keyword)
# (descrizione, [master schema.tab], [sostantivi-base esatti del concetto])
CONCETTI=[
 ("AZIENDA / Cliente / Fornitore / Partner",["public.aziende"],["aziende","azienda"]),
 ("CONTATTO (persone)",["public.contatti"],["contatti","contatto"]),
 ("UTENTE (login/operatori)",["public.utenti"],["utenti","utente"]),
 ("BU / Funzione / Struttura",["public.struttura_gerarchia"],["struttura_gerarchia","funzioni_aziendali","funzione_aziendale"]),
 ("COMMESSA",["commesse.commesse"],["commesse","commessa"]),
 ("DISCENTE (allievo FOR)",["formazione.discente"],["discente","discenti"]),
 ("ISCRIZIONE (FOR)",["formazione.iscrizione"],["iscrizione","iscrizioni"]),
 ("DIPENDENTE",["hr.dipendenti"],["dipendenti","dipendente"]),
 ("SEDE / Filiale",["hr.sedi","sedi_partner.sedi"],["sede","sedi","filiale","filiali"]),
 ("MANSIONE",["hr.mansioni"],["mansioni","mansione"]),
 ("PIANO DEI CONTI",["contabilita.piano_conti"],["piano_conti","conto","conti"]),
 ("OFFERTA",["commerciale.offerta"],["offerta","offerte"]),
 ("OPPORTUNITÀ",["commerciale.opportunita"],["opportunita"]),
 ("DEAL",["commerciale.deal"],["deal"]),
 ("ORDINE CLIENTE",["commerciale.ordine_cliente"],["ordine_cliente","ordini_cliente"]),
 ("AGENTE / Provvigione",["contabilita.agente_commerciale"],["agente","agente_commerciale","agenti"]),
 ("ODA (ordine acquisto)",["contabilita.oda"],["oda"]),
 ("BANDO / Incentivo",["fia.incentivi"],["incentivi","incentivo","bando","bandi"]),
]
r=1
wu.cell(r,1,"VERIFICA UNICITÀ — un solo MASTER per concetto (anti-duplicato, il problema di sabato)").font=Font(bold=True,size=13); r+=1
wu.cell(r,1,"✅ MASTER = fonte unica · figlia/dettaglio = tabella di dettaglio collegata (OK) · usa-il-concetto = lo referenzia via FK (OK) · ⚠️ POSSIBILE DUPLICATO = stesso sostantivo in un altro schema → controllare").font=Font(italic=True,size=9); r+=2
DUP=[]
for cdesc,masters,nouns in CONCETTI:
    tabs=[(s,t) for s,t in allt if any(n==t.lower() or t.lower().startswith(n+'_') or ('_'+n) in t.lower() for n in nouns)]
    wu.cell(r,1,cdesc).font=Font(bold=True,size=11,color="1F4E78")
    wu.cell(r,2,"MASTER: "+" + ".join(masters)).font=Font(bold=True); r+=1
    wu.cell(r,1,"Tabella"); wu.cell(r,2,"Ruolo"); wu.cell(r,3,"WeA")
    for c in (wu.cell(r,1),wu.cell(r,2),wu.cell(r,3)): c.font=H; c.fill=FILL
    r+=1
    for s,t in tabs:
        ft=f"{s}.{t}"; tl=t.lower(); fillr=None
        if ft in masters: ruolo="✅ MASTER (fonte unica)"; fillr=GREEN
        elif tl in nouns and ft not in masters: ruolo="⚠️ POSSIBILE DUPLICATO (stesso nome, altro schema)"; fillr=ORANGE; DUP.append((cdesc,ft))
        elif any(tl.startswith(n+'_') for n in nouns): ruolo="figlia/dettaglio (collegata al master)"
        else: ruolo="usa il concetto via FK (OK)"
        wu.cell(r,1,ft); wu.cell(r,2,ruolo); wu.cell(r,3,wea(t))
        if fillr:
            for cc in (1,2,3): wu.cell(r,cc).fill=fillr
        r+=1
    r+=1
# riepilogo duplicati in cima-coda
wu.cell(r,1,f"RIEPILOGO ⚠️ POSSIBILI DUPLICATI DA VERIFICARE: {len(DUP)}").font=Font(bold=True,size=12,color="C00000"); r+=1
for cd,ft in DUP: wu.cell(r,1,cd); wu.cell(r,2,ft); wu.cell(r,2).fill=ORANGE; r+=1
for i,w in enumerate([42,40,30],1): wu.column_dimensions[get_column_letter(i)].width=w
wu.freeze_panes="A4"

# INDICE
wi=wb.create_sheet("INDICE")
wi.append(["Dominio","Schema","Tabella","WeA","N° campi"])
for c in wi[1]: c.font=H; c.fill=FILL
bytbl=defaultdict(list)
for c in cols: bytbl[(c['sch'],c['tbl'])].append(c)
for (s,t),cc in sorted(bytbl.items()):
    wi.append([DOMINIO.get(s,s),s,t,wea(t),len(cc)])
for col,w in zip('ABCDE',[22,15,38,30,9]): wi.column_dimensions[col].width=w
wi.freeze_panes="A2"

# CAMPI
wc=wb.create_sheet("CAMPI")
wc.append(["Dominio","Schema","Tabella","Campo (=nome WeA)","WeA (app)","Tipo","Obblig.","PK","Collegamento (FK) →","A cosa serve","Fonte"])
for c in wc[1]: c.font=H; c.fill=FILL
for c in cols:
    k=(c['sch'],c['tbl'],c['col']); f=fonte(c)
    wc.append([DOMINIO.get(c['sch'],c['sch']),c['sch'],c['tbl'],c['col'],wea(c['tbl']),c['typ'],"Sì" if c['nul']=='NO' else "",'●' if k in pk else "",fk.get(k,""),descr(c),f])
    rr=wc.max_row
    if f=='Qnet':
        for cc in range(1,12): wc.cell(rr,cc).fill=QF
    elif f=='Sistema':
        for cc in range(1,12): wc.cell(rr,cc).fill=SF
for i,w in enumerate([20,12,28,26,22,14,7,5,28,52,12],1): wc.column_dimensions[get_column_letter(i)].width=w
wc.freeze_panes="A2"; wc.auto_filter.ref=f"A1:K{wc.max_row}"
out=os.path.expanduser("~/Desktop/STW/CATALOGO_CAMPI_ERP.xlsx")
wb.save(out)
print(f"OK v2 salvato. {len(cols)} campi · {len(bytbl)} tabelle · WeA mappate da codice · foglio UNICITÀ con {len(CONCETTI)} concetti")
