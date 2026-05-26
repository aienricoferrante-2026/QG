#!/usr/bin/env python3
"""Genera l'Excel da inviare a Ciro per la mappatura Qnet API V2.

Pre-compila tutto quello che già sappiamo (camelCase, descrizione, header
Excel, sample, coverage) e lascia in bianco le 5 colonne che SOLO CIRO può
riempire (tabella DB, nome campo SQL, tipo, endpoint API V2, note)."""
import os, sys, json
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit('Manca openpyxl. Installa con: pip3 install openpyxl')

from lib import alias_loader
from api_docs.config import BU_JSON, load_descriptions, ROOT
from api_docs import analyzer

OUT_PATH = os.path.join(ROOT, 'docs', 'qnet-schema', 'qnet-fields-template-per-ciro.xlsx')
os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)

# ── Consumatori: dove ogni campo viene usato oggi o sarà usato a breve ──
# Aiuta Ciro a capire il contesto e l'impatto di ogni campo.
WEBAPP_USERS = {
    'common':  [
        'Dashboard CM-* (11 BU singole)',
        'Dashboard Hub cross-settore (KPI consolidati tutte BU)',
        'Dashboard Totale aggregata',
        'Dashboard COGE · Contabilità Generale (legacy)',
        'WeA CdG · Controllo di Gestione',
        'WeA CON · Q-CONT Contabilità',
        'Q-WORK · tasks e attività trasversali',
        'Hub · dizionario campi',
    ],
    'FOR':     ['Dashboard CM-FOR', 'Dashboard FOR_OPP', 'Dashboard COGE',
                'Partner Views · viste dedicate per partner formazione',
                'WeA CON (pagamenti regionali)'],
    'ISO':     ['Dashboard CM-ISO', 'Dashboard COGE', 'WeA ISO · webapp dedicata'],
    'SIC':     ['Dashboard CM-SIC', 'Dashboard COGE', 'Q-SIC81 · webapp sicurezza 81/08'],
    'SOA':     ['Dashboard CM-SOA', 'Dashboard COGE'],
    'AVV':     ['Dashboard CM-AVV', 'Dashboard COGE'],
    'GAR':     ['Dashboard CM-GAR', 'Dashboard COGE'],
    'FIA':     ['Dashboard CM-FIA', 'Dashboard COGE'],
    'GDPR':    ['Dashboard CM-GDPR', 'Dashboard COGE'],
    'IST':     ['Dashboard CM-IST', 'Dashboard COGE'],
    'APL_PAL': ['Dashboard CM-APL_PAL', 'Dashboard COGE'],
    'APL_RES': ['Dashboard CM-APL_RES', 'Dashboard COGE'],
    'OFFERTE': ['Dashboard offerte', 'Dashboard Hub cross-settore',
                'WeA Sales · CRM commerciale'],
    'OPP_FOR': ['Dashboard FOR_OPP', 'WeA Sales · CRM commerciale'],
}

# ── Stili ──
HDR_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
HDR_FONT = Font(color='FFFFFF', bold=True, size=11)
TO_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # giallo per "da compilare"
DONE_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # verde per "già fatto"
BORDER = Border(left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC'))


def build_instructions_sheet(ws):
    ws.title = '📖 Istruzioni'
    ws['A1'] = 'Mappatura Qnet API V2 · campi commesse / offerte / opportunità'
    ws['A1'].font = Font(size=16, bold=True, color='1F4E78')

    text = [
        '',
        'Ciao Ciro!',
        '',
        'Per costruire l\'API che le webapp Qualifica chiameranno (Q-WORK, dashboard STW, WeA CdG, ecc.)',
        'ho bisogno di sapere, per ogni campo che noi mostriamo a video, come si chiama davvero in Qnet.',
        '',
        'Nel foglio "📋 Campi" trovi 111 righe (un campo per riga). Le prime 6 colonne sono già compilate',
        'con quello che già sappiamo dalla nostra parte. Le ultime 5 colonne (in GIALLO) sono quelle che',
        'mi puoi compilare tu.',
        '',
        '🟢 GIÀ COMPILATO (verde — non toccare):',
        '   A · BU              · es. FOR, ISO, SOA, OFFERTE, common (= comune a tutte)',
        '   B · Chiave camelCase · come la chiamiamo nelle dashboard (es. soaAttestante)',
        '   C · Header Excel IT  · come compare nell\'export Excel di Qnet (es. "Soa Attestante")',
        '   D · Descrizione      · cosa rappresenta in italiano',
        '   E · Esempio valore   · un valore reale visto nei nostri dati',
        '   F · Coverage         · su quante commesse il campo è popolato (es. "613/613")',
        '   G · Usato in         · DOVE viene consumato il campo:',
        '                          dashboard STW + webapp Qualifica (attuali e a breve)',
        '                          → ti dà il contesto del valore di ogni campo',
        '',
        '🟡 DA COMPILARE TU (giallo):',
        '   H · Tabella Qnet     · es. "orders", "ord_progress", "tasks"',
        '   I · Nome campo SQL   · es. "soa_certifier_id", "amount_consulting", "progress_pct"',
        '   J · Tipo SQL         · es. "varchar(255)", "decimal(10,2)", "datetime", "FK→users.id"',
        '   K · Endpoint API V2  · es. "GET /api/v2/orders", "GET /api/v2/orders/{id}/notes"',
        '   L · Note             · qualsiasi cosa: "FK joinata con users.full_name",',
        '                          "deprecato dal 2025", "presente solo se status=Concluso", ecc.',
        '',
        '⚠️ NB importante:',
        '   - se un campo NON ESISTE in Qnet (lo calcoliamo noi) → scrivi "CALCOLATO" nella H',
        '   - se un campo viene da JSON / dati liberi → scrivi "JSONB" e indica la struttura nella K',
        '   - se un campo è in più tabelle → scrivi quella primaria in G e le altre in K',
        '',
        '🔴 SPECIALE per la colonna K (Endpoint API V2):',
        '   - se il campo è ESPOSTO da V2 (es. GET /api/v2/orders) → scrivilo',
        '   - se il campo NON è esposto da V2 (esiste in DB ma non c\'è endpoint) → scrivi "NON IN V2"',
        '   - se per esporlo basta una piccola modifica → scrivi "DA AGGIUNGERE A V2" + indicazione',
        '   Questa colonna ci dice quanto è automatizzabile la sync vs quanto resta manuale.',
        '',
        'Tempo stimato: ~20-30 minuti. Grazie!',
        '',
        '— Enrico',
    ]
    for i, t in enumerate(text, 2):
        ws.cell(row=i, column=1).value = t
    ws.column_dimensions['A'].width = 100


def build_fields_sheet(ws, rows):
    ws.title = '📋 Campi'
    headers = [
        ('BU', 12, 'done'),
        ('Chiave camelCase', 28, 'done'),
        ('Header Excel italiano', 35, 'done'),
        ('Descrizione', 50, 'done'),
        ('Esempio valore', 30, 'done'),
        ('Coverage', 14, 'done'),
        ('Usato in (dashboard + webapp)', 45, 'done'),
        ('Tabella Qnet', 22, 'todo'),
        ('Nome campo SQL', 28, 'todo'),
        ('Tipo SQL', 24, 'todo'),
        ('Endpoint API V2', 32, 'todo'),
        ('Note Ciro', 50, 'todo'),
    ]
    for col, (label, width, _) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'A2'

    for ri, row in enumerate(rows, 2):
        for ci, (val, kind) in enumerate(zip(row, ['done']*7 + ['todo']*5), 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = BORDER
            cell.fill = DONE_FILL if kind == 'done' else TO_FILL


def collect_rows():
    """Per ogni BU genera le righe da mettere in tabella.
    Le chiavi comuni (presenti in TUTTE le BU) sono raggruppate sotto 'common'."""
    common_alias, by_bu_alias = alias_loader.load()
    desc = load_descriptions()

    # Trova chiavi comuni: presenti in >=10 BU (su 13)
    presence = {}  # key -> set di BU
    bu_fields = {}  # bu -> {key -> field_info}
    for bu in BU_JSON:
        fields, total = analyzer.analyze(bu)
        if fields is None:
            continue
        bu_fields[bu] = (fields, total)
        for k in fields:
            presence.setdefault(k, set()).add(bu)

    common_keys = {k for k, bus in presence.items() if len(bus) >= 10}

    rows = []

    # Sezione "common"
    for k in sorted(common_keys):
        sample = None; cov_total = 0; cov_nonnull = 0
        for bu, (fields, total) in bu_fields.items():
            if k in fields:
                if sample is None and fields[k]['sample'] is not None:
                    sample = fields[k]['sample']
                cov_nonnull += fields[k]['nonnull']
                cov_total += total
        excel_hdr = _italian_for(k, common_alias, by_bu_alias, None)
        users = ' · '.join(WEBAPP_USERS['common'])
        rows.append(['common', k, excel_hdr, desc.get(k, ''), _trunc(sample),
                     f'{cov_nonnull}/{cov_total} ({100*cov_nonnull//cov_total if cov_total else 0}%)',
                     users])

    # Sezioni per BU (solo campi NON comuni)
    for bu, (fields, total) in bu_fields.items():
        for k in sorted(fields):
            if k in common_keys:
                continue
            f = fields[k]
            excel_hdr = _italian_for(k, common_alias, by_bu_alias, bu)
            users = ' · '.join(WEBAPP_USERS.get(bu, [f'Dashboard CM-{bu}']))
            rows.append([bu, k, excel_hdr, desc.get(k, ''), _trunc(f['sample']),
                         f'{f["nonnull"]}/{total} ({100*f["nonnull"]//total if total else 0}%)',
                         users])

    return rows


def _italian_for(camel, common_alias, by_bu_alias, bu):
    """Reverse lookup: camelCase → header italiano (per BU + fallback comune)."""
    if bu and bu in by_bu_alias:
        for it, cc in by_bu_alias[bu].items():
            if cc == camel: return it
    for it, cc in common_alias.items():
        if cc == camel: return it
    return ''


def _trunc(v):
    if v is None: return ''
    s = str(v)
    return s[:50] + '...' if len(s) > 53 else s


def main():
    wb = openpyxl.Workbook()
    build_instructions_sheet(wb.active)
    fields_ws = wb.create_sheet()
    rows = collect_rows()
    build_fields_sheet(fields_ws, rows)
    wb.save(OUT_PATH)
    print(f'✓ Scritto {OUT_PATH}')
    print(f'  {len(rows)} campi da mappare')
    by_section = {}
    for r in rows: by_section[r[0]] = by_section.get(r[0], 0) + 1
    for sec, n in sorted(by_section.items()):
        print(f'    {sec:10s} {n} campi')


if __name__ == '__main__':
    main()
